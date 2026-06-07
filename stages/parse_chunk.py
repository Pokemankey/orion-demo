import io

import structlog
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook

import storage
from models import DocumentChunk, DocumentRef, DocumentType, Submission

log = structlog.get_logger()

# ~800 tokens expressed in characters (1 token ≈ 4 chars)
CHUNK_SIZE = 3200


def run(submission: Submission, run_id: str) -> list[DocumentChunk]:
    """
    Stage 2 — PARSE & CHUNK
    Parse each document and split into fixed-size text chunks.
    Returns a flat list of DocumentChunk objects across all documents.
    """
    all_chunks: list[DocumentChunk] = []

    for doc_ref in submission.documents:
        chunks = _parse_document(doc_ref)
        all_chunks.extend(chunks)
        log.info(
            "document_parsed",
            run_id=run_id,
            path=doc_ref.path,
            doc_type=doc_ref.type,
            chunk_count=len(chunks),
        )

    log.info("parse_complete", run_id=run_id, total_chunks=len(all_chunks))
    return all_chunks


def _parse_document(doc_ref: DocumentRef) -> list[DocumentChunk]:
    # Fetch the document's bytes from storage once (local disk in the demo,
    # object storage in production), then parse from an in-memory stream so we
    # never depend on a writable local disk — important under serverless.
    data = storage.read_bytes(doc_ref.path)
    stream = io.BytesIO(data)

    if doc_ref.type == DocumentType.PDF:
        return _parse_pdf(stream, doc_ref.path)
    elif doc_ref.type == DocumentType.DOCX:
        return _parse_docx(stream, doc_ref.path)
    elif doc_ref.type == DocumentType.XLSX:
        return _parse_xlsx(stream, doc_ref.path)
    else:
        raise ValueError(f"Unsupported document type: {doc_ref.type}")


def _parse_pdf(stream: io.BytesIO, path: str) -> list[DocumentChunk]:
    reader = PdfReader(stream)
    chunks = []

    for page_num, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if not text:
            continue

        # A single PDF page may exceed CHUNK_SIZE — split if needed
        for chunk_index, sub_chunk in enumerate(_split_text(text)):
            chunks.append(DocumentChunk(
                source_path=path,
                page=page_num,
                chunk_index=chunk_index,
                text=sub_chunk,
            ))

    return chunks


def _parse_docx(stream: io.BytesIO, path: str) -> list[DocumentChunk]:
    doc = DocxDocument(stream)
    # Collect all non-empty paragraph texts
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    full_text = "\n".join(paragraphs)

    chunks = []
    for chunk_index, text in enumerate(_split_text(full_text)):
        chunks.append(DocumentChunk(
            source_path=path,
            page=1,           # DOCX has no page concept at parse time
            chunk_index=chunk_index,
            text=text,
        ))

    return chunks


def _parse_xlsx(stream: io.BytesIO, path: str) -> list[DocumentChunk]:
    wb = load_workbook(stream, read_only=True, data_only=True)
    chunks = []
    chunk_index = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []

        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows_text.append("\t".join(cells))

        if not rows_text:
            continue

        sheet_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text)
        for text in _split_text(sheet_text):
            chunks.append(DocumentChunk(
                source_path=path,
                page=1,
                chunk_index=chunk_index,
                text=text,
            ))
            chunk_index += 1

    wb.close()
    return chunks


def _split_text(text: str) -> list[str]:
    """Split text into chunks of at most CHUNK_SIZE characters, breaking on whitespace."""
    if len(text) <= CHUNK_SIZE:
        return [text]

    chunks = []
    while text:
        if len(text) <= CHUNK_SIZE:
            chunks.append(text)
            break
        # Find the last whitespace within the limit to avoid mid-word cuts
        split_at = text.rfind(" ", 0, CHUNK_SIZE)
        if split_at == -1:
            split_at = CHUNK_SIZE
        chunks.append(text[:split_at].strip())
        text = text[split_at:].strip()

    return chunks
